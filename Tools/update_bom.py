"""Regenerate Hardware/RebotModel/BOM.xlsx with Alibaba.com sources."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

OUT = "Hardware/RebotModel/BOM.xlsx"

HEADER = ["#", "Part Name", "Specifications", "Quantity",
          "Alibaba.com Source", "Notes"]

# Each row: part_name, specs, qty, alibaba_url, notes
ROWS = [
    ("Servo", "FEETECH ST3032 serial bus servo", 2,
     "https://www.alibaba.com/trade/search?SearchText=FEETECH+ST3032+servo",
     "Search results for FEETECH ST3032 bus servo (left ID=1, right ID=2, calibrate to 2048)"),
    ("Brushless Motor", "2208, <200 KV", 2,
     "https://www.alibaba.com/trade/search?SearchText=2208+brushless+gimbal+motor+200kv",
     "2208 low-kv gimbal-style BLDC motor; pair with L6234PD013TR + AS5600"),
    ("Neodymium Magnet", "\u03d84\u00d71 mm, diametrically magnetized",
     2,
     "https://www.alibaba.com/trade/search?SearchText=4x1mm+neodymium+disc+magnet+N35",
     "For AS5600 magnetic encoder; ensure diametric (radial) magnetization"),
    ("Battery", "1S Li-Po, 1000 mAh, 3.7 V", 1,
     "https://www.alibaba.com/product-detail/603450-Rechargeable-3-7V-1000mAh-3_1601536460286.html",
     "603450 1S 3.7 V 1000 mAh Li-Po (similar form factor to original Taobao part)"),
    ("Screw M2\u00d74", "M2\u00d74 socket head cap, SS304", 20,
     "https://www.alibaba.com/product-detail/M2-Socket-Head-Cap-Machine-Screws_11000033263760.html",
     "DIN912 SHCS, multiple lengths from same supplier"),
    ("Screw M2\u00d76", "M2\u00d76 socket head cap, SS304", 10,
     "https://www.alibaba.com/product-detail/M2-Socket-Head-Cap-Machine-Screws_11000033263760.html",
     "Same SHCS family as M2\u00d74; pick the M2\u00d76 length"),
    ("Screw M2.5\u00d75", "M2.5\u00d75 socket head cap", 10,
     "https://www.alibaba.com/wholesale/round-socket-head-cap-screw.html",
     "DIN912 M2.5\u00d75 SHCS (stainless or titanium options listed)"),
    ("Screw M3\u00d76", "M3\u00d76 socket head cap, SS304", 4,
     "https://www.alibaba.com/product-detail/M3x6-M4x8-M5x10-Knurled-Socket-Head_1601777391704.html",
     "M3\u00d76 SHCS, zinc-coated stainless"),
    ("Big-head Screw M2\u00d76",
     "M2\u00d76 flat/button head (\u22485 mm head)", 4,
     "https://www.alibaba.com/product-detail/M2x6mm-Head-5mm-Black-plated-Big_1600240311485.html",
     "M2\u00d76 black flat countersunk \u22485 mm head"),
    ("Nut M2", "M2 hex nut, DIN934, SS304", 10,
     "https://www.alibaba.com/product-detail/DIN934-Nut-A2-A4-Stainless-Steel_1600459339925.html",
     "Standard DIN934 stainless hex nut"),
    ("Screw M2.3\u00d715",
     "M2.3\u00d715 self-tapping pan-head screw", 2,
     "https://www.alibaba.com/product-detail/A2-Stainless-Steel-M1-7-M2_1601015392423.html",
     "M2.3\u00d715 PT (plastic) self-tapping screw"),
    ("Bearing 3-6-2", "3\u00d76\u00d72 mm shielded miniature bearing",
     4,
     "https://www.alibaba.com/product-detail/MR-series-MR63ZZ-3x6x2-5-mm_1600892060161.html",
     "MR63ZZ family (closest catalog match; verify 2 mm width vs 2.5 mm)"),
    ("Bearing 4-7-2.5",
     "4\u00d77\u00d72.5 mm shielded miniature bearing", 4,
     "https://www.alibaba.com/product-detail/4x7x2-5-mm-Miniature-Deep-Groove_1601595369580.html",
     "MR74-2RS / MR74ZZ deep-groove ball bearing"),
    ("Cylindrical Pin",
     "\u03d84\u00d712 mm dowel pin, hardened steel", 2,
     "https://www.alibaba.com/trade/search?SearchText=4x12+dowel+pin+stainless+steel",
     "Standard \u03d84\u00d712 mm cylindrical dowel pin"),
    ("Brass Standoff M2\u00d78+3",
     "M2 brass standoff, body 8 mm + stud 3 mm", 2,
     "https://www.alibaba.com/trade/search?SearchText=M2+brass+standoff+male+female",
     "M-F brass standoff, body 8 mm + male stud 3 mm"),
    ("Brass Standoff M2\u00d725+4",
     "M2 brass standoff, body 25 mm + stud 4 mm", 2,
     "https://www.alibaba.com/trade/search?SearchText=M2+brass+standoff+PCB+spacer",
     "M-F brass standoff, body 25 mm + male stud 4 mm"),
    ("Brass Standoff M2\u00d78",
     "M2 brass female-female standoff, 8 mm", 2,
     "https://www.alibaba.com/trade/search?SearchText=M2+brass+female+standoff+spacer",
     "F-F brass standoff, 8 mm length"),
    ("Brass Standoff M2\u00d73\u00d73",
     "M2 brass round standoff, W3 \u00d7 H3 mm", 4,
     "https://www.alibaba.com/trade/search?SearchText=M2+brass+round+standoff+3mm",
     "Short M2 brass standoff (3 mm OD, 3 mm height)"),
    ("Tire", "\u03d825 \u00d7 W10 mm silicone tire, black", 2,
     "https://www.alibaba.com/trade/search?SearchText=silicone+tire+25mm+robot+wheel",
     "Silicone ring tire (25 \u00d7 10 mm) for the wheel hub"),
    ("OLED Display", "1.21\" (or 1.3\" SH1106) I\u00b2C OLED", 1,
     "https://www.alibaba.com/trade/search?SearchText=1.3+inch+OLED+SH1106+I2C+module",
     "1.21\u20131.3\" I\u00b2C OLED; the 1.3\" SH1106 module is the closest mainstream equivalent"),
    ("SH1.0 4-pin cable", "SH1.0 4P, 150 mm", 2,
     "https://www.alibaba.com/trade/search?SearchText=SH1.0+4pin+cable+150mm",
     "SH 1.0 mm pitch, 4-pin, 150 mm"),
    ("MX1.27 7-pin cable", "MX1.27 / 1.27 mm 7P, 10 mm", 1,
     "https://www.alibaba.com/trade/search?SearchText=1.27mm+pitch+7P+wire+harness",
     "1.27 mm-pitch 7-pin connector cable, ~10 mm length"),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="305496", end_color="305496", fill_type="solid"
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="center", wrap_text=True)

    ws.append(HEADER)
    for col_idx, _ in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for idx, (name, spec, qty, url, note) in enumerate(ROWS, start=1):
        ws.append([idx, name, spec, qty, url, note])
        row = ws.max_row
        link_cell = ws.cell(row=row, column=5)
        link_cell.hyperlink = url
        link_cell.style = "Hyperlink"
        for col_idx in range(1, len(HEADER) + 1):
            ws.cell(row=row, column=col_idx).alignment = wrap

    widths = {1: 5, 2: 24, 3: 38, 4: 9, 5: 70, 6: 55}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    wb.save(OUT)
    print(f"Wrote {OUT} with {len(ROWS)} rows.")


if __name__ == "__main__":
    main()
